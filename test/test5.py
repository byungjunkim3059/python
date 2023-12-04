import tkinter as tk
from tkinter import filedialog
import xlrd
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook

import pandas as pd

root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xls")])

import xlrd

def find_cell_address_in_xls(file_path, target_value):
    # .xls 파일 열기
    workbook = xlrd.open_workbook(file_path)

    # 모든 시트 순회
    for sheet_index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_index)

        # 모든 행 순회
        for row_index in range(sheet.nrows):
            # 모든 열 순회
            for col_index in range(sheet.ncols):
                cell_value = sheet.cell_value(row_index, col_index)

                # 특정 값이 발견되면 셀의 주소 반환
                if cell_value == target_value:
                    cell_address = xlrd.colname(col_index) + str(row_index + 1)
                    return cell_address

    # 특정 값이 발견되지 않은 경우 None 반환
    return None

# 사용 예제
xls_file_path = file_path  # 실제 .xls 파일 경로로 대체
target_value = " PACKING LIST  "  # 실제로 찾고자 하는 문자열로 대체

cell_address = find_cell_address_in_xls(xls_file_path, target_value)

print(cell_address)

if cell_address:
    print(f"'{target_value}'가 있는 셀의 주소: {cell_address}")
else:
    print(f"'{target_value}'를 찾을 수 없습니다.")




# def find_row_and_column_by_value(wb):
#     # 엑셀 파일 열기
#     workbook = wb
#     # 특정 시트 선택
#     sheet = workbook.active

#     # 엑셀 시트를 순회하며 특정 값이 있는 셀 찾기
#     for row in sheet.iter_rows():
#         for cell in row:
#             if cell.value == target_value:
#                 # 특정 값을 찾은 경우 행과 열 반환
#                 row_number = cell.row
#                 column_letter = openpyxl.utils.get_column_letter(cell.column)
#                 workbook.close()
#                 return row_number, column_letter

#     # 특정 값이 없는 경우 None 반환
#     workbook.close()
#     return None, None

# # 사용 예제
# target_value = "ITEM"  # 실제로 찾고자 하는 문자열로 대체

# row, column = find_row_and_column_by_value(convert_xls_to_xlsx(input_xls_path))

# if row is not None and column is not None:
#     print(f"'{target_value}'가 있는 셀의 행: {row}, 열: {column}")
# else:
#     print(f"'{target_value}'를 찾을 수 없습니다.")