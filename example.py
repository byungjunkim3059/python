import xlrd
import openpyxl
        
        
          
          

        
        
          
          # xls파일을 xlsx로 변경 
        
        
          
          def convert_xls_to_xlsx(xls_file_path):
        
        
          
              xlsBook = xlrd.open_workbook(xls_file_path)
        
        
          
              workbook = openpyxl.Workbook()
        
        
          
          

        
        
          
              for i in range(0, xlsBook.nsheets):
        
        
          
                  xlsSheet = xlsBook.sheet_by_index(i)
        
        
          
                  sheet = workbook.active if i == 0 else workbook.create_sheet()
        
        
          
                  sheet.title = xlsSheet.name
        
        
          
          

        
        
          
                  for row in range(0, xlsSheet.nrows):
        
        
          
                      for col in range(0, xlsSheet.ncols):
        
        
          
                          colvalue = xlsSheet.cell_value(row, col)
        
        
          
                          if isinstance(colvalue, str):
        
        
          
                              colvalue = colvalue.replace('', ' ', 3)		
        
        
          
          	
        
        
          
                          sheet.cell(row=row + 1, column=col + 1).value = colvalue
        
        
          
              
        
        
          
              return workbook