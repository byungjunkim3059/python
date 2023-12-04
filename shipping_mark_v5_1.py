import pandas as pd
import xlrd
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk


import win32com.client
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side

import os
import shutil
from datetime import datetime




# root.withdraw()




file_path = ""
def open_file_dialog():

    global file_path
    
    file_path = filedialog.askopenfilename(initialdir="C:\\Users\\bnj30\\Desktop\\10월 출고서류", filetypes=[("Excel Files", "*.xls")])
    print(f"Selected File: {file_path}")
    if file_path:
        label_result1.config(text=f"선택한 파일: {os.path.basename(file_path)}")
    else:
        label_result1.config(text="파일 선택이 취소되었습니다.")

file_path2 = ""
def open_file_dialog2():
    
    global file_path2

    file_path2 = filedialog.askopenfilename(initialdir="C:\\Users\\bnj30\\Desktop\\쉬핑마크\\���θ�ũ", title="쉽핑 마크 파일 선택", filetypes=[("Excel Files", "*.xlsx")])
    print(f"Selected File: {file_path}")
    if file_path2:
        label_result2.config(text=f"선택한 파일: {os.path.basename(file_path2)}")
    else:
        label_result2.config(text="파일 선택이 취소되었습니다.")
    

root = tk.Tk()
root.geometry("800x600")
root.title("쉽핑마크 시스템")

style = ttk.Style()
style.configure('TButton', padding=10, relief="flat", background="#4CAF50", foreground="black", font=('맑은 고딕', 12))


button_style = {
    'font': ('맑은 고딕', 14),  # 한글 폰트 설정
    'bg': '#4CAF50',            # 배경색 설정
    'fg': 'white',              # 글자색 설정
    'relief': 'solid',          # 테두리 스타일 설정
    'borderwidth': 2,           # 테두리 두께 설정
    'width': 20,
    'height': 1,
}
style = ttk.Style()
style.configure("BW.TLabel", foreground="black", background="white")

label_result1 = ttk.Label(text="", style="BW.TLabel")
label_result1.pack(pady=10)


button1 = ttk.Button(root, text="멋진 버튼", command=open_file_dialog)
button1.pack(side="left", pady=20, padx=20)


label_result2 = ttk.Label(text="", style="BW.TLabel")
label_result2.pack(pady=10)

button2 = tk.Button(root, text="쉽핑마크 선택하기", command=open_file_dialog2,  **button_style)
button2.pack(side="left", pady=20, padx=20)

button3 = tk.Button(root, text="프로그램 시작!", command=root.destroy)
button3.pack(pady=50, padx=20)


root.mainloop()

# file_path = filedialog.askopenfilename(initialdir = "C:\\Users\\bnj30\Desktop\\10월 출고서류", filetypes=[("Excel Files", "*.xls")])


if file_path:
    try:
        df = pd.read_excel(file_path, engine="xlrd")


        # 고객명 도출
        customer_row_indices = df[df.apply(lambda row: row.astype(str).str.contains('수         신 :').any(), axis=1)].index
        customer_columns_indices = [i for i, col in enumerate(df.columns) if df[col].astype(str).str.contains('수         신 :').any()]    

        customer_row_index = customer_row_indices[0]
        customer_column_index = customer_columns_indices[0]

        customer_info = df.iloc[customer_row_index, customer_column_index + 1]    

        # 출고일 도출
        outbound_date_row_indices = df[df.apply(lambda row: row.astype(str).str.contains('입   고   일:').any(), axis=1)].index
        outbound_date_columns_indices = [i for i, col in enumerate(df.columns) if df[col].astype(str).str.contains('입   고   일:').any()]    

        outbound_date_row_index = outbound_date_row_indices[0]
        outbound_date_column_index = outbound_date_columns_indices[0]

        outbound_date_info = df.iloc[outbound_date_row_index, outbound_date_column_index + 1]   


        # ==============================================================================================

        item_row_indices = df[df.apply(lambda row: row.astype(str).str.contains('ITEM').any(), axis=1)].index
        item_columns_indices = [i for i, col in enumerate(df.columns) if df[col].astype(str).str.contains('ITEM').any()]

        # ITEM이라는 값이 들어 있는 셀의 행과 열 주소값
        item_row_index = item_row_indices[0]
        item_column_index = item_columns_indices[0]

        item_data = df.iloc[item_row_index + 1:, item_column_index]
        item_data = item_data.dropna()

        row_start_index = item_row_index + 1
        row_end_index = row_start_index + len(item_data)

        # buyer, color, styleNo, packing, baleNo 열 주소값 구하기
        buyer_columns_indices = [i for i, col in enumerate(df.columns) if df[col].astype(str).str.contains('BUYER').any()]
        color_columns_indices = [i for i, col in enumerate(df.columns) if df[col].astype(str).str.contains('COLOR').any()]
        styleNo_columns_indices = [i for i, col in enumerate(df.columns) if df[col].astype(str).str.contains('STYLE NO').any()]
        packing_columns_indices = [i for i, col in enumerate(df.columns) if df[col].astype(str).str.contains('PACKING').any()]
        baleNo_columns_indices = [i for i, col in enumerate(df.columns) if df[col].astype(str).str.contains('Bale No.').any()]
        shippingMark_columns_indices = [i for i, col in enumerate(df.columns) if df[col].astype(str).str.contains('S/M').any()]

        totalQty_columns_indices = [i for i, col in enumerate(df.columns) if df[col].astype(str).str.contains("Total Q'ty").any()]

        buyer_column_index = buyer_columns_indices[0]
        color_column_index = color_columns_indices[0]
        styleNo_column_index = styleNo_columns_indices[0]
        packing_column_index = packing_columns_indices[0]
        baleNo_column_index = baleNo_columns_indices[0]
        shippingMark_column_index = shippingMark_columns_indices[0]
        totalQty_column_index = totalQty_columns_indices[0]

        # 6개 속성에 대해 데이터 구하기
        buyer_data = df.iloc[row_start_index:row_end_index, buyer_column_index]
        color_data = df.iloc[row_start_index:row_end_index, color_column_index]
        styleNo_data = df.iloc[row_start_index:row_end_index,styleNo_column_index]
        packing_data = df.iloc[row_start_index:row_end_index, packing_column_index]
        baleNo_data = df.iloc[row_start_index:row_end_index, baleNo_column_index]
        shippingMark_data = df.iloc[row_start_index:row_end_index, shippingMark_column_index]


        totalQty_data = df.iloc[row_start_index:row_end_index, totalQty_column_index]
        

        # baleNo 빈 곳 채우기 & 똥갈이 df 만들기
        seperate_amount_lst = []
        for k in range(len(baleNo_data)):
            i = k + row_start_index
            if pd.isna(baleNo_data[i]):
                baleNo_data[i] = (baleNo_data[i - 1])
            baleNo_data[i] = str(baleNo_data[i])

            if totalQty_data[i] % 100 != 0 and pd.notna(totalQty_data[i]):
                num = totalQty_data[i] % 100
                lst = [customer_info, item_data[i], color_data[i], num]
                seperate_amount_lst.append(lst)

        seperate_amount_columns = ['CUSTOMER', 'ITEM', 'COLOR', "Q'TY"]
        seperate_amount_df = pd.DataFrame(data=seperate_amount_lst, columns=seperate_amount_columns)

        data = []
        for k in range(len(item_data)):
            i = k + row_start_index
            lst = [buyer_data[i], item_data[i], color_data[i], styleNo_data[i], packing_data[i], baleNo_data[i], shippingMark_data[i]]
            data.append(lst)

        columns = ['BUYER', 'ITEM', 'COLOR', 'STYLE NO', 'PACKING', 'Bale No.', 'S/M']
        new_df = pd.DataFrame(data=data, columns=columns)

        # print(new_df)
        
        
        # 병합된 셀 리스트
        merged_row_index_lst = []

        # packing 열에서 빈칸인 경우 바로 앞의 데이터를 가져오는 코드
        for i in range(len(new_df['PACKING'])):
            if pd.isna(new_df['PACKING'][i]):
                new_df['PACKING'][i] = new_df['PACKING'][i - 1]
                merged_row_index_lst.append(i)
        
        merged_lst = []
        lst = []
        for i in range(len(merged_row_index_lst)):
            if i >= len(merged_row_index_lst) - 1:
                lst.append(merged_row_index_lst[i])
                merged_lst.append(lst)
                break
            lst.append(merged_row_index_lst[i])
            if merged_row_index_lst[i + 1] - merged_row_index_lst[i] > 1:
                merged_lst.append(lst)
                lst = []

        for i in range(len(merged_lst)):
            first_value = merged_lst[i][0] - 1
            merged_lst[i].append(first_value)
            merged_lst[i] = sorted(merged_lst[i])





        # 병합되지 않은 셀들의 인덱스 구하기
        merged_values = []
        for lst in merged_lst:
            for value in lst:
                merged_values.append(value)
        

        for i in range(len(item_data)):
            if i not in merged_values:
                lst = [i]
                merged_lst.append(lst)
        # 최종 셀들 병합 여부 저장된 리스트
        merged_lst = sorted(merged_lst)

        

        # 'COLOR'에 W, B를 WHITE / BLACK으로 넣어주기
        for index in range(len(new_df['COLOR'])):
            new_df['COLOR'][index] = new_df['COLOR'][index].replace("W", "WHITE")
            new_df['COLOR'][index] = new_df['COLOR'][index].replace("B", "BLACK")
            new_df['COLOR'][index] = new_df['COLOR'][index].replace("CH", "CHACOAL")

        # 'PACKING' 분해하기
        for index in range(len(new_df['PACKING'])):
            if not(pd.isna(new_df['PACKING'][index])):
                new_df['PACKING'][index] = new_df['PACKING'][index].replace(" ", "")
                new_df['PACKING'][index] = new_df['PACKING'][index].replace("\n", "")
                
                str1 = ""
                gh_str = ""
                if "(" in new_df['PACKING'][index]:
                    start_index = new_df['PACKING'][index].find('(')
                    str1= new_df['PACKING'][index][:start_index]
                    gh_str = " " + new_df['PACKING'][index][start_index:]
                else:
                    str1 = new_df['PACKING'][index]

                ct_lst = str1.split('+')

                lst = []
                for var in ct_lst:
                    x_index = var.find('x')

                    b_index = 0
                    if 'b' in var:
                        b_index = var.find('b')
                    elif 'C/T' in var:
                        b_index = var.find('C')
                    y_str = var[:x_index] + gh_str
                    bale_num = int(var[x_index + 1:b_index])

                    for i in range(bale_num):
                        lst.append(y_str)

                new_df['PACKING'][index] = lst
                

        def remove_duplicates_function(lst):
            str2 = ""
            for i in range(len(lst)):
                str2 += lst[i]
                if i >= len(lst) - 1:
                    break
                str2 += " / "
            return str2


        final_lst = []
        for lst in merged_lst:
            # 여기가 하나의 packing -> 속성들 중복 제거하는 단위!!
            buyer_lst = []
            item_lst = []
            color_lst = []
            styleNo_lst = []
            packing_lst = []
            baleNo_lst = []
            shippingMark_lst = []

            
            for index in lst:
                if new_df['BUYER'][index] not in buyer_lst:
                    buyer_lst.append(new_df['BUYER'][index])

                if new_df['ITEM'][index] not in item_lst:
                    item_lst.append(new_df['ITEM'][index])

                if new_df['COLOR'][index] not in color_lst:
                    color_lst.append(new_df['COLOR'][index])

                if new_df['STYLE NO'][index] not in styleNo_lst:
                    styleNo_lst.append(new_df['STYLE NO'][index])
                
                if new_df['PACKING'][index] != packing_lst:
                    packing_lst = new_df['PACKING'][index]

                if new_df['Bale No.'][index] not in baleNo_lst:
                    baleNo_lst.append(new_df['Bale No.'][index])
                
                if new_df['S/M'][index] not in shippingMark_lst:
                    shippingMark_lst.append(new_df['S/M'][index])


            buyer_str = remove_duplicates_function(buyer_lst)
            item_str = remove_duplicates_function(item_lst)
            color_str = remove_duplicates_function(color_lst)
            styleNo_str = remove_duplicates_function(styleNo_lst)
            baleNo_str = remove_duplicates_function(baleNo_lst)
            shippingMark_str = remove_duplicates_function(shippingMark_lst)

            
            final_lst.append([buyer_str, item_str, color_str, styleNo_str, packing_lst, baleNo_str, shippingMark_str])



        new_data = []

        for i in final_lst:
            for bale in i[4]:
                lst = [i[0], i[1], i[2], i[3], bale, i[5], i[6]]
                new_data.append(lst)


        result_df = pd.DataFrame(data=new_data, columns=columns)
        
        print(customer_info)
        year = outbound_date_info.year
        month = outbound_date_info.month
        day = outbound_date_info.day - 1

        if day < 10:
            daystr = "0" + str(day)
        else:
            daystr = str(day)

        if month < 10:
            monthstr = "0" + str(month)
        else:
            monthstr = str(month)

        folder_name = str(year) + "-" + monthstr + "-" + daystr + "_출고_" + customer_info


        
        print(result_df)
        
        rows_as_lists1 = result_df.values.tolist()
        for index, row_lst in enumerate(rows_as_lists1):
            inner_lst = [customer_info]
            for data in row_lst:
                inner_lst.append(data)
            rows_as_lists1[index] = inner_lst
        
        rows_as_lists2 = seperate_amount_df.values.tolist()

        # print(all_rows_as_one_list)

        def get_day_of_week(day_of_week):
    
            days = ["월", "화", "수", "목", "금", "토", "일"]
            day_name = days[day_of_week]

            return day_name
        
        date_object = datetime(year, month, day)
        day_of_week = date_object.weekday()

        day_name = get_day_of_week(day_of_week)

        def create_excel_table_with_columns(file_path, columns):
            # 엑셀 워크북 및 워크시트 생성
            workbook = openpyxl.Workbook()
            worksheet = workbook.active

            # 헤더(컬럼) 추가
            for col_num, column in enumerate(columns, 1):
                cell = worksheet.cell(row=1, column=col_num, value=column)
                cell.font = Font(bold=True)

            # 엑셀 파일 저장
            workbook.save(file_path)


        def write_data(wpath, data):
            try:
                workbook = openpyxl.load_workbook(wpath)
            except FileNotFoundError:
                workbook = openpyxl.Workbook()

            worksheet = workbook.active

            for row_number, row_data in enumerate(data, start=worksheet.max_row):
                row_data_with_number = [row_number] + row_data
                worksheet.append(row_data_with_number)

            last_row = worksheet.max_row
            for col in worksheet.iter_cols(min_row=last_row, max_row=last_row):
                for cell in col:
                    cell.border = Border(bottom=Side(style='thin'))

            workbook.save(wpath)

        def create_folder(path):
            try:
                columns = ['NO.', 'CUSTOMER_INFO', 'BUYER', 'ITEM', 'COLOR', 'STYLE NO', 'PACKING', 'Bale No.', 'S/M']
                # 폴더 생성
                os.makedirs(path)
                print(f"폴더가 성공적으로 생성되었습니다. 경로: {path}")
                excel_file_path1 = "C:\\Users\\bnj30\Desktop\\출고서류\\" + str(year) + "년\\" + str(month) + "월\\" + str(year) + '-' + monthstr + "-" + daystr + "(" + day_name + ")\\" + str(year) + '-' + monthstr + '-' + daystr + "(" + day_name + ")_출고.xlsx"
                write_data(excel_file_path1, rows_as_lists1)
                
                excel_file_path2 = "C:\\Users\\bnj30\Desktop\\출고서류\\" + str(year) + "년\\" + str(month) + "월\\" + str(year) + '-' + monthstr + "-" + daystr + "(" + day_name + ")\\" + str(year) + '-' + monthstr + '-' + daystr + "(" + day_name + ")_똥갈이.xlsx"
                write_data(excel_file_path2, rows_as_lists2)


                inner_excel_path = path + "\\_쉽핑마크_데이터.xlsx"
                create_excel_table_with_columns(inner_excel_path, columns)
                write_data(inner_excel_path, rows_as_lists1)
            except FileExistsError:
                print(f"폴더가 이미 존재합니다. 경로: {path}")
            except Exception as e:
                print(f"폴더 생성 중 오류가 발생했습니다. 오류 내용: {e}")

        file_name = os.path.basename(file_path).replace('.xls', '')
        folder_path = "C:\\Users\\bnj30\\Desktop\\출고서류\\" + str(year) + "년\\" + str(month) + "월\\" + str(year) + '-' + monthstr + "-" + daystr + "(" + day_name + ")\\" + file_name
        create_folder(folder_path)


        # directory_path = os.path.dirname(file_path)

        # print(directory_path)

        # if __name__ == '__main__':
        #     shutil.copytree(
        #         directory_path,
        #         folder_path,
        #         dirs_exist_ok=True
        #     )

        shutil.copy(file_path, folder_path)
    except Exception as e:
        print(f"파일 열기 오류: {str(e)}")
else:
    print("파일 선택이 취소되었습니다.")
# =======================================================================================




if file_path2:
    try:
        workbook = openpyxl.load_workbook(file_path2)

        def count_occurrences_in_excel(target_string):
            try:
                # 모든 시트 돌면서 특정 문자열 찾기
                total_occurrences = 0
                
                sheet = workbook.worksheets[0]
                occurrences_in_sheet = sum(row.count(target_string) for row in sheet.iter_rows(values_only=True))
                total_occurrences += occurrences_in_sheet
                # print(f"Sheet '{sheet_name}': {occurrences_in_sheet} occurrences")

                return total_occurrences

            except Exception as e:
                print(f"An error occurred: {e}")

        # 엑셀 파일 경로와 찾을 문자열 지정
        excel_file_path = file_path2
        search_target = 'ITEM' or "Q'TY"

        # 함수 호출
        item_num_in_sheet = count_occurrences_in_excel(search_target)

        num_rows = len(result_df['ITEM'])

        # 복사할 시트 수
        num_of_sheet = round(num_rows / item_num_in_sheet + 0.49)

        print(item_num_in_sheet)
        print(num_of_sheet)

        workbook.close()



        #Excel 프로그램 객체 생성
        excel=win32com.client.Dispatch("Excel.Application")

        #엑셀 실행과정이 보이게 설정
        excel.Visible = True

        wb1 = excel.Workbooks.Open(file_path2)

        wb2 = excel.Workbooks.Add() #엑셀 프로그램에 Workbook 추가(객체 설정)
        ws = wb2.Worksheets("sheet1") #Worksheet 설정

        for i in range(num_of_sheet):
            wb1.ActiveSheet.Copy(Before=wb2.Worksheets("sheet1"))

        ws.Delete()


        print('---')

        shipping_mark_excel_path = folder_path + "\\_쉽핑마크_완성본.xlsx"
        wb1.Save()
        print(shipping_mark_excel_path)
        wb2.SaveAs(shipping_mark_excel_path)


        excel.Quit()

        wb = openpyxl.load_workbook(shipping_mark_excel_path)

        def find_cell_address_in_all_sheets(target_value):
            # 엑셀 파일 열기

            # 모든 시트에서 특정 값 찾기
            target_cells = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    for col_idx, cell_value in enumerate(row, start=1):
                        if cell_value == target_value:
                            # 현재 셀의 주소와 행 번호를 얻어서 리스트에 추가
                            cell_address = openpyxl.utils.get_column_letter(col_idx + 1) + str(row_idx)
                            lst = [sheet_name, cell_address]
                            target_cells.append(lst)

            return target_cells

        # 예시: "existing_file.xlsx" 파일에서 모든 시트에서 값이 "찾을값"인 셀 모두 찾기
        print(find_cell_address_in_all_sheets("ITEM"))

        item_cell_address_list = find_cell_address_in_all_sheets("ITEM")
        styleNo_cell_address_list = find_cell_address_in_all_sheets("STYLE NO.") or find_cell_address_in_all_sheets("STYLE NO")
        buyer_cell_address_list = find_cell_address_in_all_sheets("BUYER") or find_cell_address_in_all_sheets("BRAND")
        color_cell_address_list = find_cell_address_in_all_sheets("COLOR")
        quantity_cell_address_list = find_cell_address_in_all_sheets("Q'TY") or find_cell_address_in_all_sheets("Q`TY") or find_cell_address_in_all_sheets("QUANTITY")
        baleNo_cell_address_list = find_cell_address_in_all_sheets("C/T NO.") or find_cell_address_in_all_sheets("C/T NO")

        wb.close()

        #Excel 프로그램 객체 생성
        excel=win32com.client.Dispatch("Excel.Application")

        #엑셀 실행과정이 보이게 설정
        excel.Visible = True

        wb = excel.Workbooks.Open(shipping_mark_excel_path)

        print(result_df["ITEM"])

        def write_data(col_lst, valuelst):
            if len(col_lst) > 0:
                for index in range(len(valuelst)):
                    ws = wb.Worksheets(col_lst[index][0])
                    ws.Range(col_lst[index][1]).Value = valuelst[index]

        
        write_data(item_cell_address_list, result_df["ITEM"])
        write_data(styleNo_cell_address_list, result_df["STYLE NO"])
        write_data(buyer_cell_address_list, result_df["BUYER"])
        write_data(color_cell_address_list, result_df["COLOR"])
        write_data(quantity_cell_address_list, result_df["PACKING"])
        write_data(baleNo_cell_address_list, result_df["Bale No."])


        wb.Save()

        excel.Quit()


        messagebox.showinfo("", "완료!")


    except Exception as e:
        print(f"파일 열기 오류: {str(e)}")
else:
    print("파일 선택이 취소되었습니다.")