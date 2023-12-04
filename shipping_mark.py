import tkinter as tk
from tkinter import filedialog

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image

import os


root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

workbook = openpyxl.load_workbook(file_path)

def count_occurrences_in_excel(target_string):
    try:
        # 모든 시트 돌면서 특정 문자열 찾기
        total_occurrences = 0
        
        sheet = workbook.worksheets[0]
        occurrences_in_sheet = sum(row.count(target_string) for row in sheet.iter_rows(values_only=True))
        total_occurrences += occurrences_in_sheet
        # print(f"Sheet '{sheet_name}': {occurrences_in_sheet} occurrences")

        return total_occurrences

    except Exception as e:
        print(f"An error occurred: {e}")

# 엑셀 파일 경로와 찾을 문자열 지정
excel_file_path = file_path
search_target = 'ITEM'

# 함수 호출
item_num_in_sheet = count_occurrences_in_excel(search_target)

num_rows = 10

# 복사할 시트 수
num_of_sheet = round(num_rows / item_num_in_sheet + 0.5)


file_name = os.path.basename(file_path)

workbook.save(file_name)

workbook.close()

