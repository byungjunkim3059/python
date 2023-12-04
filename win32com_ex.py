#win32com.client 모듈 임포트
import win32com.client
import openpyxl

import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox


messagebox.showinfo("Information", "쉽핑 마크 파일 선택")


root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename(title="쉽핑 마크 파일 선택", filetypes=[("Excel Files", "*.xlsx")])


workbook = openpyxl.load_workbook(file_path)

def count_occurrences_in_excel(target_string):
    try:
        # 모든 시트 돌면서 특정 문자열 찾기
        total_occurrences = 0
        
        sheet = workbook.worksheets[0]
        occurrences_in_sheet = sum(row.count(target_string) for row in sheet.iter_rows(values_only=True))
        total_occurrences += occurrences_in_sheet
        # print(f"Sheet '{sheet_name}': {occurrences_in_sheet} occurrences")

        return total_occurrences

    except Exception as e:
        print(f"An error occurred: {e}")

# 엑셀 파일 경로와 찾을 문자열 지정
excel_file_path = file_path
search_target = 'ITEM' or "Q'TY"

# 함수 호출
item_num_in_sheet = count_occurrences_in_excel(search_target)

num_rows = 10

# 복사할 시트 수
num_of_sheet = round(num_rows / item_num_in_sheet + 0.49)

print(item_num_in_sheet)
print(num_of_sheet)

workbook.close()



#Excel 프로그램 객체 생성
excel=win32com.client.Dispatch("Excel.Application")

#엑셀 실행과정이 보이게 설정
excel.Visible = True

wb1 = excel.Workbooks.Open(file_path)

wb2 = excel.Workbooks.Add() #엑셀 프로그램에 Workbook 추가(객체 설정)
ws = wb2.Worksheets("sheet1") #Worksheet 설정

for i in range(num_of_sheet):
    wb1.ActiveSheet.Copy(Before=wb2.Worksheets("sheet1"))

ws.Delete()

 
wb1.Save()
wb2.SaveAs(r'C:\Users\bnj30\Desktop\출고서류\data.xlsx')


excel.Quit()

wb = openpyxl.load_workbook(r'C:\Users\bnj30\Desktop\출고서류\data.xlsx')

def find_cell_address_in_all_sheets(target_value):
    # 엑셀 파일 열기

    # 모든 시트에서 특정 값 찾기
    target_cells = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            for col_idx, cell_value in enumerate(row, start=1):
                if cell_value == target_value:
                    # 현재 셀의 주소와 행 번호를 얻어서 리스트에 추가
                    cell_address = openpyxl.utils.get_column_letter(col_idx + 1) + str(row_idx)
                    lst = [sheet_name, cell_address]
                    target_cells.append(lst)

    return target_cells

# 예시: "existing_file.xlsx" 파일에서 모든 시트에서 값이 "찾을값"인 셀 모두 찾기
print(find_cell_address_in_all_sheets("ITEM"))

item_cell_address_list = find_cell_address_in_all_sheets("ITEM")
styleNo_cell_address_list = find_cell_address_in_all_sheets("STYLE NO.") or find_cell_address_in_all_sheets("STYLE NO")
buyer_cell_address_list = find_cell_address_in_all_sheets("BUYER") or find_cell_address_in_all_sheets("BRAND")
color_cell_address_list = find_cell_address_in_all_sheets("COLOR")
quantity_cell_address_list = find_cell_address_in_all_sheets("Q'TY") or find_cell_address_in_all_sheets("Q`TY") or find_cell_address_in_all_sheets("QUANTITY")
baleNo_cell_address_list = find_cell_address_in_all_sheets("C/T NO.") or find_cell_address_in_all_sheets("C/T NO")

wb.close()

#Excel 프로그램 객체 생성
excel=win32com.client.Dispatch("Excel.Application")

#엑셀 실행과정이 보이게 설정
excel.Visible = True

wb = excel.Workbooks.Open(r'C:\Users\bnj30\Desktop\출고서류\data.xlsx')

def write_data(lst, value):
    for sheet_lst in lst:
        ws = wb.Worksheets(sheet_lst[0])
        ws.Range(sheet_lst[1]).Value = value

write_data(item_cell_address_list, "S234")
write_data(styleNo_cell_address_list, "1234")
write_data(buyer_cell_address_list, "NIKE")
write_data(color_cell_address_list, "WHITE")
write_data(quantity_cell_address_list, "2000Y")
write_data(baleNo_cell_address_list, "1")


wb.Save()

excel.Quit()